from openpyxl import *

book=load_workbook(filename='julikayi.xlsx')
sheet=book['RESULTS']

print sheet.cell(row=1,column=4).value

"""#print sheet1.cell(row,column).valuefor i in range(1,rows+1	print sheet1.cell(
row=i,column=1).value


print rowsdata = [sheet.(0, 
col) forcell_value col in range(sheet.ncols)
rows=sheet1.max_rowcolumns=sheet1.max_column
print columns
	#for j in range(1,columns+1):
	#	print sheet1.cell(row=i,column=j).value
"""




"""
'c1':'c5'):

workbook = xlrd.open_workbook('me.xls')
sheet = workbook.sheet_by_index(0)
]

workbook = xlwt.Workbook()
sheet = workbook.add_sheet('test')

for index, value in enumerate(data):
    sheet.write(0, index, value)

workbook.save('me2.xls')"""